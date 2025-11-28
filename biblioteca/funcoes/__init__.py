from itertools import count
from time import sleep
def formatadata(msg):
    while True:
        data = str(input(msg)).strip()
        if data.count('/') < 2:
            print(f'Data {data} faltando "/" Redigite: ' )
        else:
            data1 = str(data).replace('/',' ').split(' ')

            dia = int(data1[0])

            mes = int(data1[1])

            ano = int(data1[2])
            if dia > 31 or mes > 12 or ano < 0 or len(data) < 10 or data.count('/') < 2:
                print('Data inválida ! ')

            else:
                return data

def verificaTexto(msg):
    while True:
        descricao = str(input(msg)).strip()
        if descricao in '' or descricao in ' ':
            print('Descrição Inválida ! Digite uma nova.')
        else:
            return descricao

def validaValor(msg):
    valor = 0
    ok = False
    while True:
        n = str(input(msg)).strip().replace(',','.')
        try:
            n = float(n)
        except Exception as error:
            print(f'Ocorreu um erro de: {error.__cause__}.')
        else:
            valor = n
            ok = True
        if ok:
            break
    return valor

def validaStatus(msg):
    while True:
        status = str(input(msg)).strip().upper()
        if status == 'DEBITO' or status == 'CREDITO':
            return status
        else:
            print('Aceita-se a palavra "CREDITO" ou "DEBITO".  Redigite ! ')

def limpar_tela():
    import os
    os.system('cls' if os.name=='nt'else 'clear')

def linha(tam=50):
    print('-'*tam)

def cabecalho(msg):
    linha()
    print(msg)
    linha()

def balanco():
    from openpyxl import Workbook, load_workbook
    from time import sleep
    extrato = [[], []]
    while True:
        limpar_tela()
        cabecalho('Escolha o tipo de operação')
        while True:
            tipo = str(input("Tipo de Operação [D ou C]: ")).strip().upper()[0]
            if tipo not in 'DC':
                print('Digite Tipo de Operação [D] ou [C]')
            else:
                break
        if tipo in 'D':
            extrato[1].append(validaValor('Digite o Valor do Débito: R$ '))
        elif tipo in 'C':
            extrato[0].append(validaValor('Digite o valor do Crédito: R$ '))
        linha()
        resp = str(input('Quer Continuar ?(S/N) ')).strip().upper()[0]
        if resp in 'Nn':
            break
    linha()
    limpar_tela()
    cabecalho('Créditos Digitados: ')
    for i, credito in enumerate(extrato[0]):
        print(f'{i} ==> R$ {credito:.2f}')
    sleep(0.3)
    cabecalho('Débitos Digitados: ')
    for j, debito in enumerate(extrato[1]):
        print(f'{j} ==> R$ {debito:.2f}')
    print('-'*50)
    sleep(0.3)
    datainicial = formatadata('Digite data (dd/mm/aaaa): ')
    if len(extrato[0]) == 0:
        entradas = 0
    else:
        entradas = sum(extrato[0])
    if len(extrato[1]) == 0:
        saidas = 0
    else:
        saidas = sum(extrato[1])
    balanco = entradas - saidas
    print(f'{entradas} -> {saidas} -> {balanco}')
    opcao = str(input('Deseja gravar na panilha ? (S/N) ')).strip().upper()[0]
    if opcao in 'Ss':
        wb = load_workbook(r"G:\Meu Drive\despesas\balanco.xlsx")
        ws = wb.active
        linha1 = ws.max_row
        ws.cell(row=linha1 + 1, column=1, value=datainicial)
        ws.cell(row=linha1 + 1, column=2, value=entradas)
        ws.cell(row=linha1 + 1, column=3, value=saidas)
        ws.cell(row=linha1 + 1, column=4, value=balanco)
        try:
            wb.save(r"G:\Meu Drive\despesas\balanco.xlsx")
        except:
            print('Erro ao gravar arquivo.')
        else:
            print('Registro adicionado com sucesso !')

    else:
        print('Registro não gravado na planilha !')
    sleep(0.5)
    limpar_tela()
    return

def despesas(msg="WILLIAM"):
    from openpyxl import Workbook, load_workbook
    limpar_tela()
    while True:
        cabecalho(f'Entre com os lançamentos de Débitos e Créditos de: {msg}. ')
        descricao = verificaTexto('Digite a descrição do lançamento: ').upper()
        data = formatadata('Digite a Data do lançamento (dd/mm/aaaa): ')
        valor = validaValor('Digite o valor do lançamento: R$ ')
        status = validaStatus('Seu lançamento é um Crédito ou um Débito ? ')
        opcao = str(input('Deseja gravar na planilha ?(S/N) ')).strip().upper()[0]
        if opcao in 'S':
            wb = load_workbook(r"G:\Meu Drive\despesas\despesas.xlsx")
            ws = wb.active
            ws = wb[msg]
            linha1 = ws.max_row
            ws.cell(row=linha1 + 1, column=1, value=data)
            ws.cell(row=linha1 + 1, column=2, value=descricao)
            ws.cell(row=linha1 + 1, column=3, value=valor)
            ws.cell(row=linha1 + 1, column=4, value=status)
            try:
                wb.save(r"G:\Meu Drive\despesas\despesas.xlsx")
            except:
                print('Erro ao gravar arquivo.')
            else:
                print('Registro adicionado com sucesso !')
        else:
            print('Registro não gravado na planilha.')
        linha()
        resp = str(input('Quer fazer novo lançamento ? (S/N) ')).strip().upper()[0]
        limpar_tela()
        if resp in 'N':
            cabecalho('Obrigado pela Colaboração !')
            sleep(0.5)
            limpar_tela()
            break
    return
